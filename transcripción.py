# encoding: utf-8

from configobj import ConfigObj
import os
import sys
import Tkinter, tkFileDialog, tkMessageBox
from Tkinter import *

from openpyxl.reader.excel import load_workbook
from openpyxl.style import Color, Fill
# Cell background color
# _cell.style.fill.fill_type = Fill.FILL_SOLID
# _cell.style.fill.start_color.index = Color.DARKRED

__version__ = "1.0"


LOCAL_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_FILE = u'transcripción.cfg'
tkinter_umlauts=['odiaeresis', 'adiaeresis', 'udiaeresis', 'Odiaeresis', 'Adiaeresis', 'Udiaeresis', 'ssharp']



class AutocompleteEntry(Tkinter.Entry):
        
        def set_completion_list(self, completion_list):
                self._completion_list = completion_list
                self._hits = []
                self._hit_index = 0
                self.position = 0
                self.bind('<KeyRelease>', self.handle_keyrelease)               

        def autocomplete(self, delta=0):
                """autocomplete the Entry, delta may be 0/1/-1 to cycle through possible hits"""
                if delta: # need to delete selection otherwise we would fix the current position
                        self.delete(self.position, Tkinter.END)
                else: # set position to end so selection starts where textentry ended
                        self.position = len(self.get())
                # collect hits
                _hits = []
                for element in self._completion_list:
                        if element.startswith(self.get().lower()):
                                _hits.append(element)
                # if we have a new hit list, keep this in mind
                if _hits != self._hits:
                        self._hit_index = 0
                        self._hits=_hits
                # only allow cycling if we are in a known hit list
                if _hits == self._hits and self._hits:
                        self._hit_index = (self._hit_index + delta) % len(self._hits)
                # now finally perform the auto completion
                if self._hits:
                        self.delete(0,Tkinter.END)
                        self.insert(0,self._hits[self._hit_index])
                        self.select_range(self.position,Tkinter.END)
                        
        def handle_keyrelease(self, event):
                """event handler for the keyrelease event on this widget"""
                if event.keysym == "BackSpace":
                        self.delete(self.index(Tkinter.INSERT), Tkinter.END) 
                        self.position = self.index(Tkinter.END)
                if event.keysym == "Left":
                        if self.position < self.index(Tkinter.END): # delete the selection
                                self.delete(self.position, Tkinter.END)
                        else:
                                self.position = self.position-1 # delete one character
                                self.delete(self.position, Tkinter.END)
                if event.keysym == "Right":
                        self.position = self.index(Tkinter.END) # go to end (no selection)
                if event.keysym == "Down":
                        self.autocomplete(1) # cycle to next hit
                if event.keysym == "Up":
                        self.autocomplete(-1) # cycle to previous hit
                # perform normal autocomplete if event is a single key or an umlaut
                if len(event.keysym) == 1 or event.keysym in tkinter_umlauts:
                        self.autocomplete()



class Transciption(Tkinter.Tk):

        def __init__(self, *args, **kwargs):
                Tkinter.Tk.__init__(self, *args, **kwargs)

                self.title(u' Gaby\'s Transcription :) :)')

                # Button X Exit
                self.protocol("WM_DELETE_WINDOW", self.exit)

                # Save
                self.bind("<Control-s>", self.save)

                # Next
                self.bind("<Next>", self.next)

                # Previous
                self.bind("<Prior>", self.previuos)

                self.init_values()
                self.config = ConfigObj(CONFIG_FILE, encoding='utf-8')
                
                # Current row
                current_frame = Frame(self)
                current_frame.pack()
                Label(current_frame, text=u'Fila actual :  ').pack(side=LEFT)
                self.var_current_row = StringVar()
                Label(current_frame, textvariable=self.var_current_row).pack()

                # Config fields
                self.add_fields()
                
                # Button options
                options_frame = Frame(self)
                options_frame.pack()
                self.boton=Button(options_frame,text="Anterior", command=self.previuos)
                self.boton.pack(side=LEFT)
                self.boton=Button(options_frame,text="Guardar", command=self.save)
                self.boton.pack(side=LEFT)
                self.boton=Button(options_frame,text="Siguiente", command=self.next)
                self.boton.pack()
                options_frame = Frame(self)
                options_frame.pack()
                self.boton=Button(options_frame,text="Primero", command=self.first)
                self.boton.pack(side=LEFT)
                self.boton=Button(options_frame,text="Cargar planilla", command=self.load)
                self.boton.pack(side=LEFT)
                self.boton=Button(options_frame,text="Último", command=self.last)
                self.boton.pack()

                self.load()

        def init_values(self):
                self.config = None
                self.current_row = None
                self.xlsx_name = None
                self.wb = None
                self.ws = None
                self.first_entry_field = None
                self.fields = dict()
                self.vars_fields = dict()
                self.fields_stay_next = list()

        def add_fields(self):
                # Frame order
                frames = list()
                for field in self.config:
                    frames.append(Frame(self))
                    frames[-1].pack()

                for field in self.config:
                        position = int(self.config[field].get('position'))
                        frame = frames[position-1]
                        frame.pack()
                        Label(frame, text=field, anchor=E, width=35).pack(fill=X, side=LEFT)
                        self.vars_fields[field] = StringVar()
                        
                        # Autocomplete
                        if 'autocomplete' in self.config[field]:
                            self.fields[field] = AutocompleteEntry(frame, textvariable=self.vars_fields[field])
                            self.fields[field].set_completion_list(self.config[field]['autocomplete'])
                        else:
                            self.fields[field] = Entry(frame, textvariable=self.vars_fields[field])
                        self.fields[field].pack()

                        # Keep first
                        if position == 1:
                            self.first_entry_field = self.fields[field]
                            self.first_entry_field.focus_set()

                        # Keep Stay Next list
                        if 'stay_next' in self.config[field]:
                            if self.config[field]['stay_next'] == 'True':
                                self.fields_stay_next.append(self.vars_fields[field])

        def clean_entries(self, stay_next=False):
                for field in self.fields:
                    var_field = self.vars_fields[field]
                    if stay_next and var_field in self.fields_stay_next:
                        continue
                    var_field.set('')

        def show_cell(self, row, stay_next=False):
                print 'Showing' , row

                self.var_current_row.set(str(row))
                # Save current worksheet (not persist)
                self.save_ws()

                self.current_row = row

                # Show on Entries
                self.clean_entries(stay_next=stay_next)
                for field in self.config:
                        column = self.config[field].get('column')
                        if column:
                                value = self.ws.cell(column + str(row)).value
                                if value:
                                        self.vars_fields[field].set(value)

                # Set focus on first field
                if self.first_entry_field:
                    self.first_entry_field.focus_set()

                print 'Showing OK'

        def load(self):
                self.xlsx_name = tkFileDialog.askopenfilename(parent=self, title='Escoge la planilla para trabajar', defaultextension='xlsx')
                print 'Openning ', self.xlsx_name
                if not self.xlsx_name: return
                self.wb = load_workbook(self.xlsx_name)
                self.ws = self.wb.worksheets[0]

                last_row = self.ws.get_highest_row()
                self.show_cell(last_row)
                print 'Openning OK'

        def save(self, dummy=None):
                self.update_autocomplete()
                self.save_ws(persist=True)
        
        def save_ws(self, persist=False):
                print 'Saving Current Row ...'
                if self.current_row:
                        # Respaldar los actuales datos si no existe archivo
                        if not self.xlsx_name:
                                current_data = dict()
                                for field in self.fields:
                                        current_data[field] = self.vars_fields[field].get()
                                self.load()
                                self.next()
                                for field in self.fields:
                                        self.vars_fields[field].set(current_data[field])

                        # Guardar nuevo dato en el XLSX
                        for field in self.config:
                                column = self.config[field].get('column')
                                if column:
                                        cell = self.ws.cell(column + str(self.current_row))
                                        cell.value = self.vars_fields[field].get()

                        # Datos permanentes # TODO: esto debe ir en el cfg
                        defaults = [{'name':u'año', 'column':u'B', 'value':'1925'},
                                    {'name':u'provincia', 'column':u'D', 'value':'santiago'},
                                    {'name':u'sexo', 'column':u'J', 'value':'h'},
                                    ]
                        for default in defaults:
                            cell = self.ws.cell(default.get('column') + str(self.current_row))
                            cell.value = default.get('value')

                        # If persist save on xlsx file
                        if persist:
                                self.wb.save(self.xlsx_name)
                print 'Saving OK'

        def update_autocomplete(self):
                # Actualiza la lista de autocompletado en los Entry
                are_changes = False
                for field in self.fields:
                        if not 'autocomplete' in self.config[field]:
                            continue 
                        text = self.fields[field].get()
                        if text and not text in self.config[field]['autocomplete']:
                                are_changes = True
                                self.config[field]['autocomplete'].append(text)
                                self.fields[field].set_completion_list(self.config[field]['autocomplete'])
                if are_changes:
                    self.config.write()


        def first(self): self.show_cell(2)

        def last(self):
                last_row = self.ws.get_highest_row()
                self.show_cell(last_row)

        def previuos(self, dummy=None):
                if self.current_row > 2:
                        self.show_cell(self.current_row-1)

        def is_current_cell_not_empty(self):
                empty = True
                for field in self.vars_fields:
                        if 'stay_next' in self.config[field]:
                            continue
                        if self.vars_fields[field].get():
                                empty = False
                                break
                return not empty

        def next(self, dummy=None):
                self.update_autocomplete()
                # Mantener N° Sección, N° Subdelegación, Comuna Subdelegación, Inscripción cuando el siguiente está vació
                if self.is_current_cell_not_empty():
                        last_row = self.ws.get_highest_row()
                        stay_next = False
                        if self.current_row == last_row:
                            stay_next = True
                        self.show_cell(self.current_row+1, stay_next=stay_next)

        def exit(self):
            if tkMessageBox.askokcancel('Salir', u'¿Estás segura que quieres salir?'):
                self.destroy()



if __name__ == '__main__':
        t = Transciption()
        t.mainloop()