# -*- coding: utf-8 -*-

from configobj import ConfigObj
from datetime import datetime
import os
import shutil
import Tkinter, tkFileDialog, tkMessageBox, tkFont

from openpyxl.reader.excel import load_workbook
# from openpyxl.style import Color, Fill
# Cell background color
# _cell.style.fill.fill_type = Fill.FILL_SOLID
# _cell.style.fill.start_color.index = Color.DARKRED

from autocomplete_entry import AutoCompleteEntry


__version__ = "1.3"


LOCAL_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_FILE = u'transcription.cfg'


class Transciption(Tkinter.Tk):

    def __init__(self, *args, **kwargs):
        
        Tkinter.Tk.__init__(self, *args, **kwargs)
        
        # Init local data
        self.init_values()
        self.config = ConfigObj(CONFIG_FILE, encoding='utf-8')
        self.interface_texts = self.config.get('interface_texts', dict())
        
        # Info
        self.title(self.config.get('title', 'Transcription'))
        self.customFont = tkFont.Font(family = self.config.get('family_font', 'Helvetica'),
            size = int(self.config.get('family_size', 16)))

        # Windows close and Press keys
        self.protocol("WM_DELETE_WINDOW", self.exit)
        self.bind("<Control-s>", self.save)
        self.bind("<Next>", self.next)
        self.bind("<Prior>", self.previuos)

        # Current row
        current_frame = Tkinter.Frame(self)
        current_frame.pack()
        Tkinter.Label(current_frame,
            text=u'%s :  ' % self.interface_texts.get('current_row', 'Current row'),
            font=self.customFont).pack(side=Tkinter.LEFT)
        self.var_current_row = Tkinter.StringVar()
        Tkinter.Label(current_frame,
            textvariable=self.var_current_row,
            font=self.customFont).pack()

        # Config fields
        self.fields = self.config.get('fields').get('to_enter', dict())
        self.permanent_fields = self.config.get('fields', dict()).get('permanent', dict())
        self.add_fields()
        
        
        # Button options
        options_frame = Tkinter.Frame(self)
        options_frame.pack()
        self.boton = Tkinter.Button(options_frame,
            text=self.interface_texts.get('previous'),
            command=self.previuos,
            font=self.customFont)
        self.boton.pack(side=Tkinter.LEFT)
        self.boton = Tkinter.Button(options_frame,
            text=self.interface_texts.get('save'),
            command=self.save,
            font=self.customFont)
        self.boton.pack(side=Tkinter.LEFT)
        self.boton = Tkinter.Button(options_frame,
            text=self.interface_texts.get('next'),
            command=self.next,
            font=self.customFont)
        self.boton.pack()

        options_frame = Tkinter.Frame(self)
        options_frame.pack()
        self.boton = Tkinter.Button(options_frame,
            text=self.interface_texts.get('first'),
            command=self.first,
            font=self.customFont)
        self.boton.pack(side=Tkinter.LEFT)
        self.boton = Tkinter.Button(options_frame,
            text=self.interface_texts.get('load_sheet'),
            command=self.load,
            font=self.customFont)
        self.boton.pack(side=Tkinter.LEFT)
        self.boton = Tkinter.Button(options_frame,
            text=self.interface_texts.get('last'),
            command=self.last,
            font=self.customFont)
        self.boton.pack()

        self.load()

    def init_values(self):
        self.config = dict()
        self.fields = dict()
        self.permanent_fields = dict()
        self.current_row = None
        self.xlsx_name = None
        self.wb = None
        self.ws = None
        self.first_entry_field = None
        self.entries = dict()
        self.vars_fields = dict()
        self.entries_stay_next = list()

    def add_fields(self):
        # Frame order
        frames = list()
        for field in self.fields:
            frames.append(Tkinter.Frame(self))
            frames[-1].pack()

        for field in self.fields:
            position = int(self.fields[field].get('position_gui'))
            frame = frames[position-1]
            frame.pack()
            Tkinter.Label(frame,
                text=field,
                anchor=Tkinter.E,
                width=35,
                font=self.customFont).pack(fill=Tkinter.X, side=Tkinter.LEFT)
            self.vars_fields[field] = Tkinter.StringVar()
            
            if 'autocomplete' in self.fields[field]:
                self.entries[field] = AutoCompleteEntry(frame,
                    textvariable=self.vars_fields[field],
                    font=self.customFont)
                self.entries[field].set_completion_list(self.fields[field]['autocomplete'])
            else:
                self.entries[field] = Tkinter.Entry(frame,
                    textvariable=self.vars_fields[field],
                    font=self.customFont)
            self.entries[field].pack()

            # Keep focus on first field
            if position == 1:
                self.first_entry_field = self.entries[field]
                self.first_entry_field.focus_set()

            # Keep Stay Next value field
            if 'stay_next' in self.fields[field]:
                if self.fields[field]['stay_next'] == 'True':
                    self.entries_stay_next.append(self.vars_fields[field])

    def clean_entries(self, stay_next=False):
        for field in self.entries:
            var_field = self.vars_fields[field]
            if stay_next and var_field in self.entries_stay_next:
                continue
            var_field.set('')

    def show_cell(self, row, stay_next=False):
        print 'Showing' , row

        self.var_current_row.set(str(row))
        self.save_ws()
        self.current_row = row

        # Show on Entries
        self.clean_entries(stay_next=stay_next)
        for field in self.fields:
            column = self.fields[field].get('column')
            if column:
                value = self.ws.cell(column + str(row)).value
                if value:
                    self.vars_fields[field].set(value)

        # Set focus on first field
        if self.first_entry_field:
            self.first_entry_field.focus_set()

    def get_backup_file_path(self):
        now = datetime.now().strftime("%Y-%m-%d %X")
        print 'Backup... ', now
        backup_dir_path = os.path.dirname(self.xlsx_name)
        backup_dir_path = os.path.join(backup_dir_path, 'backup')
        if not os.path.exists(backup_dir_path):
            os.makedirs(backup_dir_path)
        backup_file_path = "%s_%s" % (now, os.path.basename(self.xlsx_name))
        backup_file_path = os.path.join(backup_dir_path, backup_file_path)
        return backup_file_path

    def backup(self):
        backup_file_path = self.get_backup_file_path()
        shutil.copy2(self.xlsx_name, backup_file_path)
        print 'Backup... OK!' , backup_file_path

    def save_xlsx_file(self):
        print 'Saving Excel...'
        backup_file_path = self.get_backup_file_path()
        self.wb.save(backup_file_path)
        shutil.copy2(backup_file_path, self.xlsx_name)
        print 'Saving Excel... OK!' , self.xlsx_name

    def load(self):
        self.xlsx_name = tkFileDialog.askopenfilename(parent=self,
            title=self.interface_texts.get('choose_sheet'),
            defaultextension='xlsx')
        print 'Openning ', self.xlsx_name
        if not self.xlsx_name: return
        self.wb = load_workbook(self.xlsx_name)
        self.ws = self.wb.worksheets[0]

        self.backup()

        last_row = self.ws.get_highest_row()
        self.show_cell(last_row)
        print 'Openning OK'

    def save(self, dummy=None):
        self.update_autocomplete()
        self.save_ws(persist=True)
    
    def save_ws(self, persist=False):
        if self.current_row:
            # Keep current data it file doesn't exist
            if not self.xlsx_name:
                current_data = dict()
                for field in self.entries:
                    current_data[field] = self.vars_fields[field].get()
                self.load()
                self.next()
                for field in self.entries:
                    self.vars_fields[field].set(current_data[field])

            # Save current data on XLSX object
            for field in self.fields:
                column = self.fields[field].get('column')
                if column:
                    cell = self.ws.cell(column + str(self.current_row))
                    cell.value = self.vars_fields[field].get()

            # Permanent data
            for field in self.permanent_fields:
                cell = self.ws.cell(self.permanent_fields[field].get('column') + str(self.current_row))
                cell.value = self.permanent_fields[field].get('value')

            # If persist save on xlsx file
            if persist:
                    self.save_xlsx_file()

    def update_autocomplete(self):
        # Actualiza la lista de autocompletado en los Entry
        are_changes = False
        for field in self.entries:
            if not 'autocomplete' in self.fields[field]:
                continue 
            text = self.entries[field].get()
            if text and not text in self.fields[field]['autocomplete']:
                are_changes = True
                self.fields[field]['autocomplete'].append(text)
                self.entries[field].set_completion_list(self.fields[field]['autocomplete'])
        if are_changes:
            self.config.write()

    def first(self): self.show_cell(2)

    def last(self):
        last_row = self.ws.get_highest_row()
        self.show_cell(last_row)

    def previuos(self, dummy=None):
        if self.current_row > 2:
            self.show_cell(self.current_row-1)

    def is_current_cell_not_empty(self):
        empty = True
        for field in self.vars_fields:
            if 'stay_next' in self.fields[field]:
                continue
            if self.vars_fields[field].get():
                empty = False
                break
        return not empty

    def next(self, dummy=None):
        self.update_autocomplete()
        if self.is_current_cell_not_empty():
            last_row = self.ws.get_highest_row()
            stay_next = False
            if self.current_row == last_row:
                stay_next = True
            self.show_cell(self.current_row+1, stay_next=stay_next)

    def exit(self):
        if tkMessageBox.askokcancel('Close', self.interface_texts.get('close_question')):
            self.destroy()



if __name__ == '__main__':
    t = Transciption()
    t.mainloop()